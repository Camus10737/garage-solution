"""Service d'export comptable (Module 8) : résumé financier + export CSV/Excel/PDF."""
import csv
import io

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from schemas.export_comptable import ResumeComptable
from services.pdf_service import (
    BLANC,
    BLEU,
    BLEU_CLAIR,
    GRIS_LIGNE,
    _infos_garage,
    _s,
)
from services.tax_service import TPS_RATE, TVQ_RATE

LABELS_STATUT_PAIEMENT = {"non_paye": "Non payé", "partiellement_paye": "Partiellement payé", "paye": "Payé"}
LABELS_STATUT_REPARATION = {
    "en_attente": "En attente", "en_attente_piece": "En attente de pièce",
    "en_cours": "En cours", "fini": "Fini",
}


def _dans_periode(date_str: str, date_debut: str, date_fin: str) -> bool:
    jour = (date_str or "")[:10]
    return bool(jour) and date_debut <= jour <= date_fin


def _factures_periode(db, garage_id: str, date_debut: str, date_fin: str) -> list[dict]:
    docs = list(db.collection("factures").where("garage_id", "==", garage_id).stream())
    docs = [d for d in docs if _dans_periode(d.to_dict().get("date_creation", ""), date_debut, date_fin)]

    client_ids = list({d.to_dict().get("client_id") for d in docs if d.to_dict().get("client_id")})
    noms: dict[str, str] = {}
    if client_ids:
        for cdoc in db.get_all([db.collection("clients").document(cid) for cid in client_ids]):
            if cdoc.exists:
                noms[cdoc.id] = cdoc.to_dict().get("nom", "")

    factures = []
    for doc in docs:
        data = doc.to_dict()
        data["facture_id"] = doc.id
        data["client_nom"] = noms.get(data.get("client_id", ""), "")
        factures.append(data)
    factures.sort(key=lambda f: f.get("date_creation", ""))
    return factures


def _commandes_fournisseur_periode(db, garage_id: str, date_debut: str, date_fin: str) -> list[dict]:
    docs = list(db.collection("commandes_fournisseur").where("garage_id", "==", garage_id).stream())
    docs = [d for d in docs if _dans_periode(d.to_dict().get("date_commande", ""), date_debut, date_fin)]

    fournisseur_ids = list({d.to_dict().get("fournisseur_id") for d in docs if d.to_dict().get("fournisseur_id")})
    noms: dict[str, str] = {}
    if fournisseur_ids:
        for fdoc in db.get_all([db.collection("fournisseurs").document(fid) for fid in fournisseur_ids]):
            if fdoc.exists:
                noms[fdoc.id] = fdoc.to_dict().get("nom", "")

    commandes = []
    for doc in docs:
        data = doc.to_dict()
        data["commande_id"] = doc.id
        data["fournisseur_nom"] = noms.get(data.get("fournisseur_id", ""), "")
        commandes.append(data)
    commandes.sort(key=lambda c: c.get("date_commande", ""))
    return commandes


def calculer_resume(db, garage_id: str, date_debut: str, date_fin: str) -> ResumeComptable:
    factures = _factures_periode(db, garage_id, date_debut, date_fin)
    commandes = _commandes_fournisseur_periode(db, garage_id, date_debut, date_fin)

    revenus_payes = revenus_en_attente = revenus_annules = 0.0
    tps_collectee = tvq_collectee = 0.0
    montant_paye_factures = 0.0
    comptes_a_recevoir = 0.0

    for f in factures:
        total = f.get("total_facture", 0)
        if f.get("annulee"):
            revenus_annules += total
            continue
        sous_total = f.get("total_pieces", 0) + f.get("total_services", 0)
        if f.get("statut_paiement") == "paye":
            revenus_payes += total
        else:
            revenus_en_attente += total
        tps_collectee += sous_total * TPS_RATE
        tvq_collectee += sous_total * TVQ_RATE
        montant_paye_factures += f.get("montant_paye", 0)
        comptes_a_recevoir += f.get("solde_restant", 0)

    depenses_payees = depenses_dues = 0.0
    for c in commandes:
        depenses_payees += c.get("montant_paye", 0)
        depenses_dues += c.get("solde_restant", 0)

    return ResumeComptable(
        date_debut=date_debut,
        date_fin=date_fin,
        revenus_payes=round(revenus_payes, 2),
        revenus_en_attente=round(revenus_en_attente, 2),
        revenus_annules=round(revenus_annules, 2),
        tps_collectee=round(tps_collectee, 2),
        tvq_collectee=round(tvq_collectee, 2),
        taxes_totales=round(tps_collectee + tvq_collectee, 2),
        depenses_fournisseurs_payees=round(depenses_payees, 2),
        depenses_fournisseurs_dues=round(depenses_dues, 2),
        profit_net_approximatif=round(montant_paye_factures - depenses_payees, 2),
        comptes_a_recevoir=round(comptes_a_recevoir, 2),
        comptes_a_payer=round(depenses_dues, 2),
        nombre_factures=len(factures),
        nombre_commandes_fournisseur=len(commandes),
    )


_ENTETE_FACTURES = [
    "Numéro", "Date", "Client", "Sous-total", "TPS", "TVQ", "Total",
    "Statut paiement", "Montant payé", "Solde restant", "Statut réparation", "Annulée",
]
_ENTETE_COMMANDES = ["Numéro", "Date", "Fournisseur", "Montant total", "Montant payé", "Solde dû", "Statut"]


def _ligne_facture(f: dict) -> list:
    sous_total = f.get("total_pieces", 0) + f.get("total_services", 0)
    return [
        f.get("numero_facture", f.get("facture_id", "")[:8]),
        (f.get("date_creation", ""))[:10],
        f.get("client_nom", ""),
        round(sous_total, 2),
        round(sous_total * TPS_RATE, 2),
        round(sous_total * TVQ_RATE, 2),
        f.get("total_facture", 0),
        LABELS_STATUT_PAIEMENT.get(f.get("statut_paiement"), ""),
        f.get("montant_paye", 0),
        f.get("solde_restant", 0),
        LABELS_STATUT_REPARATION.get(f.get("statut_reparation"), ""),
        "Oui" if f.get("annulee") else "Non",
    ]


def _ligne_commande(c: dict) -> list:
    return [
        c.get("numero_commande", c.get("commande_id", "")[:8]),
        (c.get("date_commande", ""))[:10],
        c.get("fournisseur_nom", ""),
        c.get("montant_total", 0),
        c.get("montant_paye", 0),
        c.get("solde_restant", 0),
        LABELS_STATUT_PAIEMENT.get(c.get("statut_paiement"), ""),
    ]


def generer_csv(factures: list[dict], commandes: list[dict]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(["Factures"])
    writer.writerow(_ENTETE_FACTURES)
    for f in factures:
        writer.writerow(_ligne_facture(f))

    writer.writerow([])
    writer.writerow(["Commandes fournisseur"])
    writer.writerow(_ENTETE_COMMANDES)
    for c in commandes:
        writer.writerow(_ligne_commande(c))

    return buffer.getvalue().encode("utf-8-sig")


def generer_excel(resume: ResumeComptable, factures: list[dict], commandes: list[dict]) -> bytes:
    wb = Workbook()

    ws_resume = wb.active
    ws_resume.title = "Résumé"
    lignes_resume = [
        ("Période", f"{resume.date_debut} au {resume.date_fin}"),
        ("Revenus payés", resume.revenus_payes),
        ("Revenus en attente", resume.revenus_en_attente),
        ("Revenus annulés", resume.revenus_annules),
        ("TPS collectée", resume.tps_collectee),
        ("TVQ collectée", resume.tvq_collectee),
        ("Taxes totales", resume.taxes_totales),
        ("Dépenses fournisseurs payées", resume.depenses_fournisseurs_payees),
        ("Dépenses fournisseurs dues", resume.depenses_fournisseurs_dues),
        ("Profit net approximatif", resume.profit_net_approximatif),
        ("Comptes à recevoir", resume.comptes_a_recevoir),
        ("Comptes à payer", resume.comptes_a_payer),
        ("Nombre de factures", resume.nombre_factures),
        ("Nombre de commandes fournisseur", resume.nombre_commandes_fournisseur),
    ]
    for row in lignes_resume:
        ws_resume.append(row)

    ws_factures = wb.create_sheet("Factures")
    ws_factures.append(_ENTETE_FACTURES)
    for f in factures:
        ws_factures.append(_ligne_facture(f))

    ws_commandes = wb.create_sheet("Fournisseurs")
    ws_commandes.append(_ENTETE_COMMANDES)
    for c in commandes:
        ws_commandes.append(_ligne_commande(c))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _table_donnees(entete: list, lignes: list, col_widths: list, formats_montant: set[int]) -> Table:
    s = _s()
    rows = [[Paragraph(h, s["th"]) for h in entete]]
    for ligne in lignes:
        row = []
        for i, val in enumerate(ligne):
            texte = f"{val:.2f} $" if i in formats_montant and isinstance(val, (int, float)) else str(val)
            row.append(Paragraph(texte, s["td_right"] if i in formats_montant else s["td"]))
        rows.append(row)

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLEU),
        ("TEXTCOLOR", (0, 0), (-1, 0), BLANC),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BLANC, colors.HexColor("#f8fafc")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, GRIS_LIGNE),
    ]))
    return table


def generer_pdf(resume: ResumeComptable, factures: list[dict], commandes: list[dict], garage_id: str) -> bytes:
    garage_nom, garage_adresse, garage_tel = _infos_garage(garage_id)
    buffer = io.BytesIO()
    page_w, page_h = letter
    margin = 0.6 * inch
    usable_w = page_w - 2 * margin

    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=margin, leftMargin=margin, topMargin=margin, bottomMargin=margin,
    )
    s = _s()
    elements = []

    # ── En-tête ──────────────────────────────────────────────────────────
    garage_col = [
        Paragraph(garage_nom, s["garage_nom"]),
        Spacer(1, 4),
        Paragraph(garage_adresse, s["garage_info"]),
        Paragraph(garage_tel, s["garage_info"]),
    ]
    titre_box = Table(
        [[Paragraph("RAPPORT COMPTABLE", s["doc_titre"])],
         [Paragraph(f"{resume.date_debut} au {resume.date_fin}", s["box_valeur"])]],
        colWidths=[2.6 * inch],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLEU),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, 1), 6),
            ("BOX", (0, 0), (-1, -1), 1, BLEU),
        ]),
    )
    header_table = Table(
        [[garage_col, titre_box]],
        colWidths=[usable_w - 2.7 * inch, 2.7 * inch],
        style=TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (1, 0), (1, 0), "RIGHT")]),
    )
    elements.append(header_table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=BLEU, spaceAfter=0.2 * inch))

    # ── Résumé ───────────────────────────────────────────────────────────
    lignes_resume = [
        ("Revenus payés", resume.revenus_payes), ("Revenus en attente", resume.revenus_en_attente),
        ("Revenus annulés", resume.revenus_annules), ("TPS collectée", resume.tps_collectee),
        ("TVQ collectée", resume.tvq_collectee), ("Taxes totales", resume.taxes_totales),
        ("Dépenses fournisseurs payées", resume.depenses_fournisseurs_payees),
        ("Dépenses fournisseurs dues", resume.depenses_fournisseurs_dues),
        ("Profit net approximatif", resume.profit_net_approximatif),
        ("Comptes à recevoir", resume.comptes_a_recevoir),
        ("Comptes à payer", resume.comptes_a_payer),
    ]
    moitie = (len(lignes_resume) + 1) // 2
    col1, col2 = lignes_resume[:moitie], lignes_resume[moitie:]

    def _bloc(col):
        return Table(
            [[Paragraph(label, s["sous_label"]), Paragraph(f"{valeur:.2f} $", s["sous_valeur"])] for label, valeur in col],
            colWidths=[2.0 * inch, 1.1 * inch],
            style=TableStyle([("ALIGN", (0, 0), (0, -1), "LEFT"), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]),
        )

    resume_table = Table(
        [[_bloc(col1), _bloc(col2)]],
        colWidths=[usable_w / 2, usable_w / 2],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), BLEU_CLAIR),
            ("BOX", (0, 0), (-1, -1), 0.5, GRIS_LIGNE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]),
    )
    elements.append(resume_table)
    elements.append(Spacer(1, 0.3 * inch))

    # ── Détail factures ──────────────────────────────────────────────────
    elements.append(Paragraph(f"Factures ({resume.nombre_factures})", s["garage_nom"]))
    elements.append(Spacer(1, 0.1 * inch))
    if factures:
        col_widths = [usable_w * w for w in (0.11, 0.09, 0.16, 0.11, 0.08, 0.08, 0.11, 0.13, 0.13)]
        entete = ["Numéro", "Date", "Client", "Sous-total", "TPS", "TVQ", "Total", "Paiement", "Solde"]
        lignes = [
            [f.get("numero_facture", f.get("facture_id", "")[:8]), (f.get("date_creation", ""))[:10], f.get("client_nom", ""),
             round(f.get("total_pieces", 0) + f.get("total_services", 0), 2),
             round((f.get("total_pieces", 0) + f.get("total_services", 0)) * TPS_RATE, 2),
             round((f.get("total_pieces", 0) + f.get("total_services", 0)) * TVQ_RATE, 2),
             f.get("total_facture", 0), LABELS_STATUT_PAIEMENT.get(f.get("statut_paiement"), ""), f.get("solde_restant", 0)]
            for f in factures
        ]
        elements.append(_table_donnees(entete, lignes, col_widths, formats_montant={3, 4, 5, 6, 8}))
    else:
        elements.append(Paragraph("Aucune facture pour cette période.", s["garage_info"]))
    elements.append(Spacer(1, 0.3 * inch))

    # ── Détail commandes fournisseur ─────────────────────────────────────
    elements.append(Paragraph(f"Commandes fournisseur ({resume.nombre_commandes_fournisseur})", s["garage_nom"]))
    elements.append(Spacer(1, 0.1 * inch))
    if commandes:
        col_widths = [usable_w * w for w in (0.15, 0.13, 0.27, 0.15, 0.15, 0.15)]
        entete = ["Numéro", "Date", "Fournisseur", "Total", "Payé", "Dû"]
        lignes = [
            [c.get("numero_commande", c.get("commande_id", "")[:8]), (c.get("date_commande", ""))[:10],
             c.get("fournisseur_nom", ""), c.get("montant_total", 0), c.get("montant_paye", 0), c.get("solde_restant", 0)]
            for c in commandes
        ]
        elements.append(_table_donnees(entete, lignes, col_widths, formats_montant={3, 4, 5}))
    else:
        elements.append(Paragraph("Aucune commande fournisseur pour cette période.", s["garage_info"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
